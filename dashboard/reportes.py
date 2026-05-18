from django.http import HttpResponse
from datetime import date
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reservas.models import Reserva


def filtrar_reservas(fecha_inicio, fecha_fin):
    """
    Filtra las reservas por rango de fechas.
    Si no se especifica fecha_inicio o fecha_fin, no filtra ese extremo.
    """
    reservas = Reserva.objects.select_related('espacio', 'usuario').all()
    
    if fecha_inicio:
        reservas = reservas.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        reservas = reservas.filter(fecha__lte=fecha_fin)
    
    return reservas.order_by('-fecha')


def generar_excel_reservas(fecha_inicio=None, fecha_fin=None):
    """
    Genera un archivo Excel (.xlsx) con las reservas filtradas.
    
    Columnas del Excel:
    ID | Espacio | Código | Usuario | Email | Fecha | Hora inicio | Hora fin | Estado | Motivo
    
    Características:
    - Encabezados con fondo azul y texto blanco
    - Ancho de columnas automático
    - Nombre del archivo: reservas_YYYY-MM-DD.xlsx
    """
    # Obtener datos filtrados
    reservas = filtrar_reservas(fecha_inicio, fecha_fin)
    
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reservas'
    
    # Definir encabezados
    headers = [
        'ID', 'Espacio', 'Código', 'Usuario', 'Email',
        'Fecha', 'Hora inicio', 'Hora fin', 'Estado', 'Motivo'
    ]
    
    # Estilos para encabezados
    header_font = Font(bold=True, color='FFFFFF')           # Texto blanco negrita
    header_fill = PatternFill('solid', fgColor='0D6EFD')    # Fondo azul
    
    # Escribir encabezados con estilos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Escribir datos de reservas
    for row_idx, reserva in enumerate(reservas, 2):
        ws.cell(row=row_idx, column=1, value=reserva.id)
        ws.cell(row=row_idx, column=2, value=reserva.espacio.nombre)
        ws.cell(row=row_idx, column=3, value=reserva.espacio.codigo)
        ws.cell(row=row_idx, column=4, value=reserva.usuario.get_full_name() or reserva.usuario.username)
        ws.cell(row=row_idx, column=5, value=reserva.usuario.email)
        ws.cell(row=row_idx, column=6, value=reserva.fecha)
        ws.cell(row=row_idx, column=7, value=str(reserva.hora_inicio))
        ws.cell(row=row_idx, column=8, value=str(reserva.hora_fin))
        ws.cell(row=row_idx, column=9, value=reserva.get_estado_display())
        ws.cell(row=row_idx, column=10, value=reserva.motivo)
    
    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reservas_{date.today()}.xlsx"'
    
    # Guardar el libro en la respuesta
    wb.save(response)
    return response


def generar_pdf_reservas(fecha_inicio=None, fecha_fin=None):
    """
    Genera un archivo PDF con las reservas filtradas.
    
    Características:
    - Orientación horizontal (landscape)
    - Tamaño carta (letter)
    - Tabla con colores alternados
    - Encabezados en azul con texto blanco
    - Nombre del archivo: reservas_YYYY-MM-DD.pdf
    """
    # Obtener datos filtrados
    reservas = filtrar_reservas(fecha_inicio, fecha_fin)
    
    # Crear buffer en memoria para el PDF
    buffer = io.BytesIO()
    
    # Configurar documento
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter)  # Horizontal, tamaño carta
    )
    
    elementos = []
    styles = getSampleStyleSheet()
    
    # Título del reporte
    elementos.append(Paragraph('Reporte de Reservas', styles['Title']))
    
    # Mostrar período si se especificó
    if fecha_inicio or fecha_fin:
        periodo_texto = f'Periodo: {fecha_inicio or "Inicio"} a {fecha_fin or "Hoy"}'
        elementos.append(Paragraph(periodo_texto, styles['Normal']))
    
    # Total de reservas
    elementos.append(Paragraph(f'Total: {reservas.count()} reservas', styles['Normal']))
    elementos.append(Spacer(1, 20))  # Espacio vertical
    
    # Datos de la tabla
    data = [['ID', 'Espacio', 'Usuario', 'Fecha', 'Inicio', 'Fin', 'Estado', 'Motivo']]
    
    for reserva in reservas:
        data.append([
            str(reserva.id),
            reserva.espacio.codigo,
            (reserva.usuario.get_full_name() or reserva.usuario.username)[:20],  # Truncar nombre
            str(reserva.fecha),
            str(reserva.hora_inicio),
            str(reserva.hora_fin),
            reserva.get_estado_display(),
            reserva.motivo[:30],  # Truncar motivo
        ])
    
    # Crear tabla con estilos
    tabla = Table(data, repeatRows=1)  # repeatRows=1: repetir encabezados en cada página
    tabla.setStyle(TableStyle([
        # Encabezados
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),  # Fondo azul
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),                  # Texto blanco
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),               # Negrita
        ('FONTSIZE', (0, 0), (-1, 0), 10),                             # Tamaño fuente
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),                           # Alineación izquierda
        
        # Datos
        ('FONTSIZE', (0, 1), (-1, -1), 8),                             # Fuente más pequeña
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),                 # Líneas de cuadrícula
        
        # Colores alternados
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.white,
            colors.HexColor('#F8F9FA')
        ]),
    ]))
    
    elementos.append(tabla)
    
    # Construir PDF
    doc.build(elementos)
    
    # Obtener PDF del buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    # Preparar respuesta HTTP
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reservas_{date.today()}.pdf"'
    
    return response